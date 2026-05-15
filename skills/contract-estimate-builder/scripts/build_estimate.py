#!/usr/bin/env python3
"""Contract Estimate Builder.

Generates an Excel bid sheet + PDF scope of work from a JSON spec.

CRITICAL: Reads brand identity (DRE, name, brokerage) from
shared-references/identity.json - never hardcoded. This is the rule across all
of Graeham's skills: identity.json is the single source of truth.
"""

import json
import re
import sys
from datetime import date
from itertools import product
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, KeepTogether, ListFlowable, ListItem, PageBreak,
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# ---- Identity (read from canonical source) ----

def load_identity():
    """Read brand identity from shared-references/identity.json.

    Walks up from this script's location to find Skills/skills/shared-references/.
    Raises a clear error if not found - we never want to fall back to hardcoded
    values because that's exactly how the zombie DRE keeps reappearing.
    """
    here = Path(__file__).resolve().parent
    # contract-estimate-builder/scripts -> contract-estimate-builder -> skills
    skills_root = here.parent.parent
    identity_path = skills_root / "shared-references" / "identity.json"
    if not identity_path.exists():
        raise FileNotFoundError(
            "Could not find identity.json at " + str(identity_path) +
            ". This skill must run from inside Skills/skills/contract-estimate-builder/. "
            "Do NOT hardcode identity values - fix the path instead."
        )
    with open(identity_path) as f:
        data = json.load(f)
    ident = data["identity"]
    blocked = data.get("_blocked_values", {}).get("dre_blocklist", [])
    if ident["dre"] in blocked:
        raise ValueError(
            "identity.json has a blocked DRE (" + ident["dre"] +
            "). Stop and fix identity.json before proceeding."
        )
    return ident


IDENTITY = load_identity()

NAVY = "1A365D"
TEAL = "0D9488"
SLATE = "475569"
LIGHT_GRAY = "F1F5F9"
ROW_ALT = "F8FAFC"
GREEN_FILL = "DCFCE7"
BLUE_INPUT = "0000FF"
DISCLAIMER_GRAY = "64748B"
DISCLAIMER_DARK = "334155"

CURRENCY_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def agent_credit():
    """Build the prepared-by signature line from identity.json values."""
    return (
        IDENTITY["name"] + ", " + IDENTITY["title"] +
        " - " + IDENTITY["brokerage"] +
        " - DRE# " + IDENTITY["dre"]
    )


DISCLAIMER_TITLE = "DISCLAIMER AND OWNER RESPONSIBILITIES"

DISCLAIMER_INTRO = (
    "This document is provided solely as a courtesy by " + IDENTITY["name"] +
    " (" + IDENTITY["title"] + ", " + IDENTITY["brokerage"] +
    ", DRE# " + IDENTITY["dre"] + ") to summarize a scope of work being "
    "discussed between the property owner and a third-party contractor. It is "
    "not a contract, not a binding offer, and not a recommendation. No signature "
    "is required, and receipt or review of this document by any party does not "
    "constitute acceptance of, or agreement to, any terms set forth in it. Any "
    "agreement for work to be performed must be entered into separately and in "
    "writing directly between the property owner and the contractor."
)

DISCLAIMER_LIABILITY = (
    IDENTITY["name"] + " and " + IDENTITY["brokerage"] + " are not a party to "
    "any agreement between the property owner and the contractor, are not "
    "performing any of the work described, and receive no referral fee, "
    "compensation, or financial benefit of any kind from the contractor. No "
    "warranty - express or implied - is made as to pricing, scope, quality of "
    "work, contractor licensure, contractor insurance, contractor bond status, "
    "contractor qualifications, materials, timeline, code compliance, permit "
    "requirements, or workmanship. " + IDENTITY["name"] + " and " +
    IDENTITY["brokerage"] + " assume no liability for performance or non-"
    "performance of the work, for any damage to person or property, or for any "
    "dispute that may arise between the owner and the contractor."
)

DISCLAIMER_LICENSING = (
    "Licensed or unlicensed contractor - owner accepts all risk. The contractor "
    "named in this scope may or may not be a licensed contractor. " + IDENTITY["name"] + " "
    "makes no representation either way and has not verified the contractor's "
    "license status. The property owner accepts all risk and responsibility for "
    "their choice of contractor. If the contractor IS licensed, the owner is "
    "responsible for verifying that license themselves at cslb.ca.gov; " +
    IDENTITY["name"] + " takes no responsibility for whether that license is "
    "current, valid, or in good standing. If the contractor is NOT licensed "
    "and the owner chooses to work with them anyway, that decision and all "
    "consequences are the owner's alone. We always recommend obtaining "
    "alternative bids if the owner does not feel comfortable with this "
    "contractor for any reason."
)


DISCLAIMER_OWNER_LEAD = "The property owner is solely responsible for:"

DISCLAIMER_OWNER_ITEMS = [
    "Verifying the contractor's current license status and bond at cslb.ca.gov "
    "(California Contractors State License Board), or knowingly accepting an "
    "unlicensed contractor at the owner's own risk. Under California Business "
    "and Professions Code Section 7048, unlicensed contractors may only perform "
    "work where the combined cost of labor and materials is less than $500.",

    "Confirming the contractor's general liability insurance and workers' "
    "compensation coverage before any work begins.",

    "Pulling any required building permits and confirming the work complies with "
    "all local building codes, HOA covenants, conditions, and restrictions, "
    "easements, and zoning restrictions.",

    "Investigating any environmental factors that may apply, including but not "
    "limited to lead paint (homes built before 1978), asbestos, mold, and any "
    "other hazardous-material considerations.",

    "Obtaining additional bids if desired and determining for themselves whether "
    "the contractor's pricing is reasonable for the work described.",

    "Inspecting completed work, accepting or rejecting completion, and resolving "
    "any disputes directly with the contractor.",

    "All other investigation, due diligence, follow-up, and research relating to "
    "the contractor, the scope, and the work to be performed. " +
    IDENTITY["name"] + " is not conducting any such investigation or due "
    "diligence on behalf of the owner.",

    "Consulting their own independent legal counsel before signing any contract "
    "or making any payment.",
]

DISCLAIMER_ACK = (
    "By receiving this document, the property owner and contractor each "
    "acknowledge that this is an informational scope summary only, that it "
    "does not constitute a contract or agreement between any parties, and that "
    "no party is bound by anything contained in it until and unless a separate "
    "written agreement is signed."
)


def slugify(s):
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")[:60] or "estimate"


def fmt_date(d):
    return d or date.today().isoformat()


def excel_disclaimer_text():
    parts = [DISCLAIMER_TITLE, "", DISCLAIMER_INTRO, "",
             DISCLAIMER_LIABILITY, "", DISCLAIMER_LICENSING, "",
             DISCLAIMER_OWNER_LEAD]
    for item in DISCLAIMER_OWNER_ITEMS:
        parts.append("- " + item)
    parts.append("")
    parts.append(DISCLAIMER_ACK)
    return "\n".join(parts)


def build_excel(spec, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Sheet"

    ws["A1"] = "CONTRACT ESTIMATE - " + spec.get("trade", "Scope of Work").upper()
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Property: " + spec["property_address"]
    ws["A3"] = "Date: " + fmt_date(spec.get("date"))
    ws["A4"] = "Contractor: " + (spec.get("contractor_name") or "____________________________")
    ws["A5"] = "Client: " + (spec.get("client_name") or "____________________________")
    for r in range(2, 6):
        ws["A" + str(r)].font = Font(size=10, color=SLATE)
        ws.merge_cells("A" + str(r) + ":F" + str(r))

    header_row = 7
    for col_idx, h in enumerate(["#", "Task", "Notes", "Qty", "Unit Cost", "Line Total"], start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    row = header_row + 1
    base_first = row
    for i, item in enumerate(spec.get("base_items", []), start=1):
        unit_cost = item.get("unit_cost")
        cells = [
            i, item.get("task", ""), item.get("notes", "") or "",
            item.get("qty", 1),
            unit_cost if unit_cost is not None else None,
            "=D" + str(row) + "*E" + str(row),
        ]
        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = BORDER
            if col_idx == 1: c.alignment = Alignment(horizontal="center")
            elif col_idx in (2, 3): c.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx in (4, 5, 6): c.alignment = Alignment(horizontal="right")
            if col_idx in (5, 6): c.number_format = CURRENCY_FMT
            if col_idx == 5: c.font = Font(color=BLUE_INPUT)
            if row % 2 == 0: c.fill = PatternFill("solid", start_color=ROW_ALT)
        row += 1
    base_last = row - 1

    base_total_row = row
    ws.cell(row=row, column=2, value="BASE TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6,
            value="=SUM(F" + str(base_first) + ":F" + str(base_last) + ")").number_format = CURRENCY_FMT
    for col_idx in range(1, 7):
        c = ws.cell(row=row, column=col_idx)
        c.fill = PatternFill("solid", start_color=LIGHT_GRAY)
        c.font = Font(bold=True)
        c.border = BORDER
    row += 2

    option_group_rows = []
    for g_idx, group in enumerate(spec.get("option_groups", []), start=1):
        ws.cell(row=row, column=1, value="OPTIONS - Group " + str(g_idx))
        ws.cell(row=row, column=2, value=group.get("label", ""))
        ws.cell(row=row, column=3, value=group.get("notes", "Pick one."))
        for col_idx in range(1, 7):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill("solid", start_color=NAVY)
            c.font = Font(bold=True, color="FFFFFF")
            c.border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

        rows_for_group = []
        for o_idx, opt in enumerate(group.get("options", []), start=1):
            unit_cost = opt.get("unit_cost")
            cells = [
                str(g_idx) + "." + str(o_idx), opt.get("name", ""),
                opt.get("notes", "") or "", opt.get("qty", 1),
                unit_cost if unit_cost is not None else None,
                "=D" + str(row) + "*E" + str(row),
            ]
            for col_idx, val in enumerate(cells, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.border = BORDER
                if col_idx == 1: c.alignment = Alignment(horizontal="center")
                elif col_idx in (2, 3): c.alignment = Alignment(wrap_text=True, vertical="top")
                elif col_idx in (4, 5, 6): c.alignment = Alignment(horizontal="right")
                if col_idx in (5, 6): c.number_format = CURRENCY_FMT
                if col_idx == 5: c.font = Font(color=BLUE_INPUT)
                c.fill = PatternFill("solid", start_color=GREEN_FILL if row % 2 == 0 else "ECFCCB")
            rows_for_group.append((opt.get("name", "Option " + str(o_idx)), row))
            row += 1
        option_group_rows.append({"label": group.get("label", "Group " + str(g_idx)), "options": rows_for_group})
        row += 1

    for col, w in {"A": 6, "B": 36, "C": 38, "D": 8, "E": 14, "F": 16}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    s2 = wb.create_sheet("Totals Summary")
    s2["A1"] = "TOTALS SUMMARY - Grand Total by Scenario"
    s2["A1"].font = Font(bold=True, size=14, color=NAVY)
    s2.merge_cells("A1:C1")
    s2["A2"] = "Property: " + spec["property_address"]
    s2["A2"].font = Font(size=10, color=SLATE)
    s2.merge_cells("A2:C2")

    s2["A4"] = "Scenario"; s2["B4"] = "Components"; s2["C4"] = "Grand Total"
    for col_idx in range(1, 4):
        c = s2.cell(row=4, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    s_row = 5
    s2.cell(row=s_row, column=1, value="Base only (no options)")
    s2.cell(row=s_row, column=2, value="Base Total")
    s2.cell(row=s_row, column=3, value="='Bid Sheet'!F" + str(base_total_row)).number_format = CURRENCY_FMT
    for col_idx in range(1, 4):
        s2.cell(row=s_row, column=col_idx).border = BORDER
    s_row += 1

    if option_group_rows:
        combos = list(product(*[g["options"] for g in option_group_rows]))
        if len(combos) <= 12:
            for combo in combos:
                labels = " + ".join(n for n, _ in combo)
                comps = "Base + " + " + ".join(n for n, _ in combo)
                parts = ["'Bid Sheet'!F" + str(base_total_row)] + ["'Bid Sheet'!F" + str(rn) for _, rn in combo]
                formula = "=" + "+".join(parts)
                s2.cell(row=s_row, column=1, value="Base + " + labels)
                s2.cell(row=s_row, column=2, value=comps)
                s2.cell(row=s_row, column=3, value=formula).number_format = CURRENCY_FMT
                for col_idx in range(1, 4):
                    s2.cell(row=s_row, column=col_idx).border = BORDER
                    if s_row % 2 == 0:
                        s2.cell(row=s_row, column=col_idx).fill = PatternFill("solid", start_color=ROW_ALT)
                s_row += 1
        else:
            for g in option_group_rows:
                s2.cell(row=s_row, column=1, value="-- " + g["label"] + " options --").font = Font(bold=True, italic=True)
                s_row += 1
                for n, rn in g["options"]:
                    s2.cell(row=s_row, column=1, value="Base + " + n)
                    s2.cell(row=s_row, column=2, value="Base + " + n)
                    s2.cell(row=s_row, column=3, value="='Bid Sheet'!F" + str(base_total_row) + "+'Bid Sheet'!F" + str(rn)).number_format = CURRENCY_FMT
                    for col_idx in range(1, 4):
                        s2.cell(row=s_row, column=col_idx).border = BORDER
                    s_row += 1

    for r in range(5, s_row):
        s2.cell(row=r, column=3).font = Font(bold=True)
    s2.column_dimensions["A"].width = 42
    s2.column_dimensions["B"].width = 48
    s2.column_dimensions["C"].width = 18
    s2.freeze_panes = s2.cell(row=5, column=1)

    note_row = s_row + 2
    s2.cell(row=note_row, column=1, value="How to use this:").font = Font(bold=True)
    s2.cell(row=note_row + 1, column=1, value="- Contractor fills in Unit Cost on the Bid Sheet tab (blue cells).").font = Font(size=10, color=SLATE)
    s2.cell(row=note_row + 2, column=1, value="- Grand Total for each scenario above updates automatically.").font = Font(size=10, color=SLATE)
    s2.cell(row=note_row + 3, column=1, value="- Options are mutually exclusive - pick ONE per group.").font = Font(size=10, color=SLATE)
    for r in range(note_row + 1, note_row + 4):
        s2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    disc_row = note_row + 5
    c = s2.cell(row=disc_row, column=1, value=excel_disclaimer_text())
    c.font = Font(size=8, color=DISCLAIMER_DARK, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    s2.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=3)
    s2.row_dimensions[disc_row].height = 320

    wb.save(out_path)


def build_pdf(spec, out_path):
    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        leftMargin=0.7 * inch, rightMargin=0.7 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        title="Contract Estimate - " + spec["property_address"],
        author=IDENTITY["name"] + ", " + IDENTITY["brokerage"],
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor("#" + NAVY), fontSize=18, leading=22, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#" + NAVY), fontSize=13, leading=16, spaceBefore=14, spaceAfter=6)
    meta = ParagraphStyle("meta", parent=styles["Normal"], textColor=colors.HexColor("#" + SLATE), fontSize=10, leading=13)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14)
    task_title = ParagraphStyle("task_title", parent=styles["Normal"], fontSize=10.5, leading=14, fontName="Helvetica-Bold")
    task_notes = ParagraphStyle("task_notes", parent=styles["Normal"], fontSize=9.5, leading=12, textColor=colors.HexColor("#" + SLATE))
    option_label = ParagraphStyle("opt", parent=styles["Normal"], fontSize=11, leading=14, fontName="Helvetica-Bold", textColor=colors.HexColor("#" + TEAL))
    disc_h = ParagraphStyle("disc_h", parent=styles["Normal"], fontSize=10, leading=13, fontName="Helvetica-Bold", textColor=colors.HexColor("#" + DISCLAIMER_DARK), spaceBefore=10, spaceAfter=4)
    disc_body = ParagraphStyle("disc_body", parent=styles["Normal"], fontSize=8, leading=11, textColor=colors.HexColor("#" + DISCLAIMER_DARK), alignment=4, spaceAfter=6)
    disc_bullet = ParagraphStyle("disc_bullet", parent=disc_body, fontSize=8, leading=11, leftIndent=12, spaceAfter=3)
    disc_ack = ParagraphStyle("disc_ack", parent=disc_body, fontSize=8, leading=11, fontName="Helvetica-Oblique", textColor=colors.HexColor("#" + DISCLAIMER_DARK))

    story = []
    story.append(Paragraph("CONTRACT ESTIMATE - " + spec.get("trade", "Scope of Work").upper(), h1))
    story.append(Paragraph("<b>Property:</b> " + spec["property_address"], meta))
    story.append(Paragraph("<b>Date:</b> " + fmt_date(spec.get("date")), meta))
    contractor = spec.get("contractor_name") or "____________________________"
    client = spec.get("client_name") or "____________________________"
    story.append(Paragraph("<b>Contractor:</b> " + contractor, meta))
    story.append(Paragraph("<b>Client:</b> " + client, meta))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=0.6, color=colors.HexColor("#" + NAVY)))
    story.append(Spacer(1, 4))

    instr = Table([[Paragraph(
        "<b>Instructions:</b> Please return pricing per line item on the attached Excel bid sheet. "
        "Items below are the proposed base scope. Where alternative options are listed, "
        "price each option separately - the client will pick one. The Excel auto-calculates "
        "a grand total for each option scenario. No signature is required on this document - "
        "it is a scope summary, not a contract.",
        body)]], colWidths=[7.0 * inch])
    instr.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + LIGHT_GRAY)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#" + SLATE)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(instr)

    story.append(Paragraph("Base Scope of Work", h2))
    base_data = [["#", "Task", "Notes", "Unit Price"]]
    for i, item in enumerate(spec.get("base_items", []), start=1):
        base_data.append([
            str(i),
            Paragraph(item.get("task", ""), task_title),
            Paragraph(item.get("notes", "") or "", task_notes),
            "$ _______________" if item.get("unit_cost") is None else "$" + format(item.get("unit_cost"), ",.2f"),
        ])
    bt = Table(base_data, colWidths=[0.4 * inch, 2.6 * inch, 2.8 * inch, 1.4 * inch], repeatRows=1)
    bt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9.5),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + ROW_ALT)]),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#" + SLATE)),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(bt)

    if spec.get("option_groups"):
        story.append(Paragraph("Alternative Options (Pick One Per Group)", h2))
        for g_idx, group in enumerate(spec["option_groups"], start=1):
            block = [Paragraph("Group " + str(g_idx) + ": " + group.get("label", ""), option_label)]
            if group.get("notes"):
                block.append(Paragraph(group["notes"], task_notes))
            block.append(Spacer(1, 4))
            opt_data = [["#", "Option", "Notes", "Unit Price"]]
            for o_idx, opt in enumerate(group.get("options", []), start=1):
                opt_data.append([
                    str(g_idx) + "." + str(o_idx),
                    Paragraph(opt.get("name", ""), task_title),
                    Paragraph(opt.get("notes", "") or "", task_notes),
                    "$ _______________" if opt.get("unit_cost") is None else "$" + format(opt.get("unit_cost"), ",.2f"),
                ])
            ot = Table(opt_data, colWidths=[0.5 * inch, 2.5 * inch, 2.8 * inch, 1.4 * inch], repeatRows=1)
            ot.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + TEAL)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9.5),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + GREEN_FILL)]),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#" + TEAL)),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A7F3D0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            block.append(ot)
            block.append(Spacer(1, 8))
            story.append(KeepTogether(block))

    story.append(PageBreak())
    story.append(Paragraph(DISCLAIMER_TITLE, disc_h))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#" + DISCLAIMER_GRAY)))
    story.append(Spacer(1, 6))
    story.append(Paragraph(DISCLAIMER_INTRO, disc_body))
    story.append(Paragraph(DISCLAIMER_LIABILITY, disc_body))
    story.append(Paragraph(DISCLAIMER_LICENSING, disc_body))
    story.append(Paragraph("<b>" + DISCLAIMER_OWNER_LEAD + "</b>", disc_body))
    bullets = [ListItem(Paragraph(item, disc_bullet), leftIndent=12, bulletColor=colors.HexColor("#" + DISCLAIMER_DARK))
               for item in DISCLAIMER_OWNER_ITEMS]
    story.append(ListFlowable(bullets, bulletType="bullet", start="circle", leftIndent=14, bulletFontSize=7))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<i>" + DISCLAIMER_ACK + "</i>", disc_ack))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Optional - Acknowledgment of receipt (not required, not a contract):", disc_body))
    sig = Table([
        ["Property owner signature", "Date"],
        ["_________________________________", "______________"],
        ["", ""],
        ["Contractor signature", "Date"],
        ["_________________________________", "______________"],
    ], colWidths=[4.5 * inch, 2.5 * inch])
    sig.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#" + SLATE)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(sig)
    story.append(Spacer(1, 14))
    story.append(Paragraph("<i>Prepared by " + agent_credit() + "</i>",
                           ParagraphStyle("footer", parent=styles["Normal"], fontSize=8,
                                          textColor=colors.HexColor("#" + SLATE), alignment=1)))
    doc.build(story)


def main():
    if len(sys.argv) < 3:
        print("Usage: python build_estimate.py <spec.json> <output_dir>")
        sys.exit(1)
    spec = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)
    slug = slugify(spec["property_address"].split(",")[0])
    xlsx = out_dir / (slug + "-estimate.xlsx")
    pdf = out_dir / (slug + "-estimate.pdf")
    build_excel(spec, xlsx)
    build_pdf(spec, pdf)
    print("Wrote:", xlsx)
    print("Wrote:", pdf)


if __name__ == "__main__":
    main()
