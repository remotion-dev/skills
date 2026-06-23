#!/usr/bin/env python3
"""
weekly-listing-update — report generator

Usage (from skill):
    python3 generate_report.py \\
        --showing-file path/to/showing.xlsx \\
        --address "1908 Cooley Ave" \\
        --city "East Palo Alto" \\
        --state CA \\
        --beds 3 --baths 1 --sqft 1210 \\
        --list-price 1095000 \\
        --mls ML82027334 \\
        --seller-name "Michael" \\
        --listed-date 2025-11-14 \\
        --output path/to/output.html

Designed to be invoked by Claude inside the weekly-listing-update skill.
The skill orchestrates: read inputs, run this script, humanize the prose,
push to GitHub via direct git, create the Gmail draft.

WEEK-OVER-WEEK MODEL (added 2026-06-22)
---------------------------------------
The uploaded ShowingTime/Glide export is CUMULATIVE: it holds every showing and
disclosure pull since the listing went live, each stamped with a date. That single
file IS the running history. This script buckets every dated row into calendar
weeks (Mon-Sun) so the report can show the climb week 1 -> week 2 -> ... -> total
and emphasize the most recent week. No separate database is needed; prior published
reports serve only as a backup cross-check.

If a row has no parseable date, it falls into an "undated" bucket and is still
counted in the cumulative totals, so the report never loses an interaction — it
just cannot place it on the weekly curve.
"""

import argparse, calendar, json, re, sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def initial_format(name):
    if not name:
        return ""
    name = str(name).strip().replace('\n', ' ').replace('  ', ' ')
    if '/' in name:
        name = name.split('/')[0].strip()
    parts = [p for p in name.split() if p and not p.startswith('(')]
    if len(parts) >= 2:
        return f"{parts[0].title()} {parts[-1][0].upper()}."
    return parts[0].title() if parts else name


def clean_feedback(fb):
    if not fb:
        return ""
    fb = str(fb).strip().replace('\n', ' ').replace('  ', ' ')
    fb = re.sub(r'^\d{1,2}/\d{1,2}/\d{2,4}[\s\-:]*', '', fb)
    return fb


def count_mentions(text, keywords):
    text = text.lower()
    return sum(text.count(k.lower()) for k in keywords)


# ---------------------------------------------------------------------------
# Date helpers — used to bucket activity into weeks
# ---------------------------------------------------------------------------

def parse_dt(val):
    """Best-effort parse of a cell value into a date. Returns datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    # strip a trailing time if it confuses the parser
    s_main = s.split(' ')[0] if ' ' in s and '/' in s.split(' ')[0] or '-' in s.split(' ')[0] else s
    fmts = ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%m-%d-%Y', '%m-%d-%y',
            '%m/%d/%Y %H:%M', '%m/%d/%y %I:%M %p', '%Y-%m-%d %H:%M:%S']
    for f in fmts:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    for f in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
        try:
            return datetime.strptime(s_main, f)
        except ValueError:
            continue
    # last resort: pull the first date-looking token
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', s)
    if m:
        mo, da, yr = m.groups()
        yr = int(yr)
        if yr < 100:
            yr += 2000
        try:
            return datetime(yr, int(mo), int(da))
        except ValueError:
            return None
    return None


def week_start(dt):
    """Monday of the week containing dt (date-only)."""
    d = datetime(dt.year, dt.month, dt.day)
    return d - timedelta(days=d.weekday())


def week_label(monday):
    return f"{calendar.month_abbr[monday.month]} {monday.day}"


NAME_HEADERS = ('agent name', 'showing agent', 'name', 'agent')
ACTION_DATE_HEADERS = ('email', 'call', 'text')  # these columns hold action timestamps
LEAD_DATE_RE = re.compile(r'^\s*(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?')


def header_row(ws):
    raw = next(ws.iter_rows(values_only=True, max_row=1), None) or ()
    return [(str(c).strip().lower() if c is not None else '') for c in raw]


def find_col(headers, candidates):
    """Exact header match first, then a 'contains' match (never the email-address column)."""
    for i, h in enumerate(headers):
        if h in candidates:
            return i
    for i, h in enumerate(headers):
        if h != 'email address' and any(c in h for c in candidates):
            return i
    return None


def is_date_header(h):
    return ('date' in h) or (h in ACTION_DATE_HEADERS)


def row_date(row, date_idxs, text_idxs, default_year):
    """Best available date for a row: the earliest real date found in any date-ish
    column, falling back to a leading M/D in the feedback/notes text. Handles both
    sheets with an explicit Date column and sheets where the only dates are the
    email/call/text action stamps (as in the hand-kept tracker)."""
    cands = []
    for i in date_idxs:
        if i is not None and i < len(row):
            d = parse_dt(row[i])
            if d:
                cands.append(d)
    for i in text_idxs:
        if i is not None and i < len(row) and row[i]:
            m = LEAD_DATE_RE.match(str(row[i]))
            if m:
                mo, da, yr = m.groups()
                year = (int(yr) + 2000 if yr and int(yr) < 100 else int(yr)) if yr else default_year
                try:
                    cands.append(datetime(year, int(mo), int(da)))
                except ValueError:
                    pass
    return min(cands) if cands else None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_sheet(ws, kind, default_year, dedupe=False):
    """Header-driven parse of one activity sheet. Locates the name, feedback,
    notes, and date columns by their headers, so it works whether feedback is in
    column 5, 6, or 10 and whether the dates are in a Date column or the
    email/call/text action columns. Returns a list of records."""
    if ws is None:
        return []
    headers = header_row(ws)
    name_idx = find_col(headers, NAME_HEADERS)
    if name_idx is None:
        return []
    fb_idx = find_col(headers, ('feedback',))
    notes_idx = find_col(headers, ('notes',))
    date_idxs = [i for i, h in enumerate(headers) if is_date_header(h)]
    text_idxs = [i for i in (fb_idx, notes_idx) if i is not None]

    records, seen = [], set()
    for row in ws.iter_rows(values_only=True, min_row=2):
        if name_idx >= len(row) or not row[name_idx]:
            continue
        name_raw = str(row[name_idx]).strip().replace('\n', ' ')
        if not name_raw or name_raw.lower().startswith('invited'):
            continue
        if dedupe:
            key = name_raw.lower()
            if key in seen:
                continue
            seen.add(key)
        fb = clean_feedback(row[fb_idx]) if (fb_idx is not None and fb_idx < len(row)) else ''
        note = clean_feedback(row[notes_idx]) if (notes_idx is not None and notes_idx < len(row)) else ''
        records.append({
            'name': initial_format(name_raw),
            'feedback': fb,
            'note': note,
            'date': row_date(row, date_idxs, text_idxs, default_year),
            'kind': kind,
        })
    return records


def parse_showing_file(path, default_year=2026):
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet(name):
        return wb[name] if name in wb.sheetnames else None

    supra = parse_sheet(sheet('Showing Activity (Suprashowing)'), 'showing', default_year)
    glide = parse_sheet(sheet('Glide View Activity'), 'disclosure', default_year, dedupe=True)
    openh = parse_sheet(sheet('Open House Feedback'), 'openhouse', default_year)

    all_records = supra + glide + openh
    all_feedback = ' '.join([(a['feedback'] + ' ' + a['note']).lower() for a in all_records])
    themes = build_themes(all_feedback)

    return {
        'supra_agents': supra,
        'glide_agents': glide,
        'openhouse_agents': openh,
        'all_records': all_records,
        'counts': {
            'showings': len(supra),
            'disclosures': len(glide),
            'open_house': len(openh),
            'total_interactions': len(all_records),
        },
        'themes': themes,
    }


def build_themes(all_feedback):
    return {
        'tenant': count_mentions(all_feedback, ['tenant', 'occupied']),
        'price': count_mentions(all_feedback, ['price', 'negotiat', 'asking']),
        'neighbor': count_mentions(all_feedback, ['neighbor']),
        'location': count_mentions(all_feedback, ['location', 'east side', 'west side', 'surrounding']),
        'no_longer': count_mentions(all_feedback, ['no longer interested', 'not interested']),
        'lost_to_other': count_mentions(all_feedback, ['purchased', 'in contract', 'went into contract', 'closed deal', 'another property', 'different property']),
        'forgot': count_mentions(all_feedback, ['forgot', 'long time ago', 'doesnt remember', "doesn't remember"]),
    }


THEME_LABELS = {
    'tenant': 'Tenant occupancy',
    'price': 'Price / negotiation interest',
    'neighbor': 'Neighbor / street concern',
    'location': 'Location preference',
    'no_longer': 'No longer interested',
    'lost_to_other': 'Lost to competing inventory',
    'forgot': 'Forgot the property',
}


# ---------------------------------------------------------------------------
# Week-over-week bucketing
# ---------------------------------------------------------------------------

def build_weekly(all_records, listed_date, report_date, recent_weeks=6):
    """Bucket records into weeks counted from the listing date: Week 1 is the
    first 7 days after listing, Week N covers days [(N-1)*7, N*7). The week that
    contains report_date is the current 'this week'. Returns ordered weekly rows
    plus this-week deltas, cumulative totals, and a momentum read."""
    days_live = max(0, (report_date - listed_date).days)
    current_idx = days_live // 7

    buckets = {}
    undated = {'showings': 0, 'disclosures': 0, 'new_feedback': 0}
    for r in all_records:
        if not r['date']:
            undated['showings'] += 1 if r['kind'] == 'showing' else 0
            undated['disclosures'] += 1 if r['kind'] == 'disclosure' else 0
            undated['new_feedback'] += 1 if r['feedback'] else 0
            continue
        idx = (r['date'] - listed_date).days // 7
        idx = max(0, min(idx, current_idx))
        b = buckets.setdefault(idx, {'showings': 0, 'disclosures': 0, 'new_feedback': 0, 'feedback': []})
        b['showings'] += 1 if r['kind'] == 'showing' else 0
        b['disclosures'] += 1 if r['kind'] == 'disclosure' else 0
        if r['feedback']:
            b['new_feedback'] += 1
            b['feedback'].append(r['feedback'].lower())

    def range_label(idx):
        start = listed_date + timedelta(days=idx * 7)
        end = start + timedelta(days=6)
        return f"{calendar.month_abbr[start.month]} {start.day}–{calendar.month_abbr[end.month]} {end.day}"

    # show every week that has activity, plus the current week even if empty.
    # Renumber from the first ACTIVE week so "Week 1" is the launch / first-activity
    # week — never an empty pre-marketing week before showings and pulls began.
    active_idxs = sorted(set(buckets) | {current_idx})
    weeks = []
    for pos, idx in enumerate(active_idxs):
        b = buckets.get(idx, {'showings': 0, 'disclosures': 0, 'new_feedback': 0, 'feedback': []})
        wk_themes = build_themes(' '.join(b['feedback']))
        top = max(wk_themes.items(), key=lambda kv: kv[1]) if any(wk_themes.values()) else (None, 0)
        weeks.append({
            'week_num': pos + 1,
            'label': f"Week {pos + 1}",
            'range': range_label(idx),
            'showings': b['showings'],
            'disclosures': b['disclosures'],
            'new_feedback': b['new_feedback'],
            'interactions': b['showings'] + b['disclosures'],
            'top_theme': THEME_LABELS.get(top[0]) if top[1] else None,
            'is_current': idx == current_idx,
        })

    this_week = next((w for w in weeks if w['is_current']), weeks[-1] if weeks else None)
    prior_week = None
    for w in weeks:
        if w is this_week:
            break
        prior_week = w

    def delta(cur, prev, key):
        if prev is None or cur is None:
            return None
        return cur[key] - prev[key]

    this_week_deltas = None
    if this_week:
        this_week_deltas = {
            'showings': delta(this_week, prior_week, 'showings'),
            'disclosures': delta(this_week, prior_week, 'disclosures'),
            'interactions': delta(this_week, prior_week, 'interactions'),
        }

    # momentum: current week vs the average of up to 3 prior weeks
    momentum = 'steady'
    priors = [w for w in weeks if w is not this_week]
    if priors and this_week:
        trail = [w['interactions'] for w in priors[-3:]]
        avg = sum(trail) / len(trail) if trail else 0
        if this_week['interactions'] > avg * 1.25 and this_week['interactions'] > 0:
            momentum = 'accelerating'
        elif this_week['interactions'] < avg * 0.6:
            momentum = 'cooling'

    cumulative = {
        'showings': sum(w['showings'] for w in weeks) + undated['showings'],
        'disclosures': sum(w['disclosures'] for w in weeks) + undated['disclosures'],
    }
    cumulative['interactions'] = cumulative['showings'] + cumulative['disclosures']

    detailed = weeks[-recent_weeks:]
    older = weeks[:-recent_weeks]
    earlier_rollup = None
    if older:
        earlier_rollup = {
            'label': f"Weeks 1–{older[-1]['week_num']}",
            'showings': sum(w['showings'] for w in older),
            'disclosures': sum(w['disclosures'] for w in older),
            'new_feedback': sum(w['new_feedback'] for w in older),
            'interactions': sum(w['interactions'] for w in older),
        }

    return {
        'weeks': weeks,
        'detailed_weeks': detailed,
        'earlier_rollup': earlier_rollup,
        'undated': undated if (undated['showings'] or undated['disclosures']) else None,
        'this_week': this_week,
        'prior_week': prior_week,
        'this_week_deltas': this_week_deltas,
        'momentum': momentum,
        'cumulative': cumulative,
        'weeks_live': current_idx + 1,
        'dated': bool(buckets),
    }


# ---------------------------------------------------------------------------
# Trajectory chart (SVG) — emitted ready-to-embed so the report never hand-builds coords
# ---------------------------------------------------------------------------

def build_chart_svg(weeks):
    """Inline SVG line chart of CUMULATIVE disclosure downloads (navy area+line)
    and showings (gold line) across the weeks, with a Start=0 origin so the line
    climbs. The skill embeds this string directly — no coordinate math in the LLM.
    Returns '' if there is nothing to plot."""
    if not weeks:
        return ""
    cum_d, cum_s, td, ts = [], [], 0, 0
    for w in weeks:
        td += w['disclosures']; ts += w['showings']
        cum_d.append(td); cum_s.append(ts)
    n = len(weeks)
    pts_n = n + 1  # include the Start=0 origin point
    W, H, L, R, T, B = 640, 240, 44, 18, 20, 36
    pw, ph = W - L - R, H - T - B
    ymax = max(4, ((cum_d[-1] + 3) // 4) * 4)
    FN = "DM Sans, system-ui, sans-serif"

    def X(i):
        return L + (pw * i / (pts_n - 1) if pts_n > 1 else 0)

    def Y(v):
        return T + ph * (1 - v / ymax)

    d_pts = [(X(0), Y(0))] + [(X(i + 1), Y(cum_d[i])) for i in range(n)]
    s_pts = [(X(0), Y(0))] + [(X(i + 1), Y(cum_s[i])) for i in range(n)]
    poly = lambda pts: " ".join("%.0f,%.0f" % (x, y) for x, y in pts)
    area = ("M%.0f,%.0f " % (X(0), Y(0))
            + " ".join("L%.0f,%.0f" % (x, y) for x, y in d_pts[1:])
            + " L%.0f,%.0f Z" % (X(n), Y(0)))

    p = ['<svg viewBox="0 0 %d %d" width="100%%" role="img" '
         'aria-label="Cumulative disclosure downloads and showings by week" '
         'style="display:block;margin:10px 0 4px;">' % (W, H),
         '<title>Cumulative disclosure downloads and showings since launch</title>']
    for v in (0, ymax // 2, ymax):
        y = Y(v)
        col = "#c5cee0" if v == 0 else "#e6e9ef"
        p.append('<line x1="%d" y1="%.0f" x2="%d" y2="%.0f" stroke="%s"/>' % (L, y, W - R, y, col))
        p.append('<text x="%d" y="%.0f" text-anchor="end" font-family="%s" font-size="11" fill="#9aa3b5">%d</text>' % (L - 8, y + 4, FN, v))
    p.append('<path d="%s" fill="#e7ebf2"/>' % area)
    p.append('<polyline points="%s" fill="none" stroke="#0f1729" stroke-width="2.5"/>' % poly(d_pts))
    p.append('<polyline points="%s" fill="none" stroke="#d49019" stroke-width="2.5"/>' % poly(s_pts))
    for i in range(n):
        dx, dy = d_pts[i + 1]; sx, sy = s_pts[i + 1]
        p.append('<circle cx="%.0f" cy="%.0f" r="4" fill="#0f1729"/>' % (dx, dy))
        p.append('<circle cx="%.0f" cy="%.0f" r="4" fill="#d49019"/>' % (sx, sy))
        p.append('<text x="%.0f" y="%.0f" text-anchor="middle" font-family="%s" font-size="12" font-weight="600" fill="#0f1729">%d</text>' % (dx, dy - 9, FN, cum_d[i]))
    labels = ["Start"] + ["Wk %d" % w['week_num'] for w in weeks]
    for i, lab in enumerate(labels):
        x = X(i)
        anc = "start" if i == 0 else ("end" if i == pts_n - 1 else "middle")
        cur = i >= 1 and weeks[i - 1].get('is_current')
        fill = "#0f1729" if cur else "#6a7488"
        fw = ' font-weight="600"' if cur else ''
        txt = lab + (" · now" if cur else "")
        p.append('<text x="%.0f" y="%d" text-anchor="%s" font-family="%s" font-size="11.5" fill="%s"%s>%s</text>' % (x, H - 14, anc, FN, fill, fw, txt))
    p.append('<rect x="%d" y="2" width="11" height="11" rx="2" fill="#0f1729"/>' % L)
    p.append('<text x="%d" y="12" font-family="%s" font-size="11.5" fill="#3a4154">Disclosure downloads (cumulative)</text>' % (L + 16, FN))
    p.append('<rect x="%d" y="2" width="11" height="11" rx="2" fill="#d49019"/>' % (L + 262))
    p.append('<text x="%d" y="12" font-family="%s" font-size="11.5" fill="#3a4154">Showings (cumulative)</text>' % (L + 278, FN))
    p.append('</svg>')
    return "".join(p)


# ---------------------------------------------------------------------------
# Offers + warm leads
# ---------------------------------------------------------------------------

OFFER_KEYWORDS = ['accepted offer', 'offer accepted', 'submitted an offer', 'submitted offer',
                  'wrote an offer', 'made an offer', 'received an offer', 'offer received',
                  'offer in', 'will write', 'writing an offer', 'in escrow']


def detect_offer_signals(data):
    """Conservative scan for offer-relevant rows. These are CANDIDATES for the
    report's offer banner — Claude confirms and enriches (price, terms, status)
    from context or from the --offers argument. Deliberately excludes
    'in contract with ANOTHER property', which means a buyer lost elsewhere."""
    signals = []
    for agent in data['all_records']:
        fb = agent['feedback'].lower()
        if not fb:
            continue
        if any(k in fb for k in OFFER_KEYWORDS):
            # exclude buyers who went under contract elsewhere
            if 'another' in fb or 'different property' in fb or 'elsewhere' in fb:
                continue
            signals.append({'name': agent['name'], 'feedback': agent['feedback'],
                            'date': agent['date'].strftime('%Y-%m-%d') if agent['date'] else None})
    return signals


def detect_warm_leads(data):
    leads = []
    keywords = {
        'price-negotiation': ['interested but', 'wants to negotiate', 'negotiate price', 'room to negotiate'],
        'active-disclosure': ['accessed disclosures', 'discuss with my clients', 'will get back'],
        'tenant-investor': ['how many tenants', 'month to month', 'how much is the rent', 'is it still occupied'],
    }
    for agent in data['all_records']:
        fb_lower = agent['feedback'].lower()
        for signal, kws in keywords.items():
            if any(k in fb_lower for k in kws):
                leads.append({'name': agent['name'], 'signal': signal})
                break
    return leads


OFFER_STATE_RANK = {'received': 1, 'countered': 2, 'accepted': 3, 'pending': 4, 'in_contract': 4}


def normalize_state(s):
    return str(s or 'received').lower().strip().replace(' ', '_')


def top_offer_state(offers):
    """Highest state across all offers. received < countered < accepted < pending/in_contract."""
    if not offers:
        return None
    best, best_rank = 'received', 0
    for o in offers:
        s = normalize_state(o.get('status'))
        r = OFFER_STATE_RANK.get(s, 1)
        if r >= best_rank:
            best_rank, best = r, ('in_contract' if s == 'in_contract' else s)
    return best


def determine_status(data, weekly, offers):
    """Offer state drives status. A RECEIVED offer is amber (a decision is due),
    NOT green. Only accepted is green; pending / in contract is blue."""
    state = top_offer_state(offers)
    if state in ('pending', 'in_contract'):
        return {'level': 'blue', 'headline': 'Status · Pending, In Contract',
                'message': 'Accepted and in escrow. Contingencies are being worked toward close.'}
    if state == 'accepted':
        return {'level': 'green', 'headline': 'Status · Offer Accepted',
                'message': 'An offer has been accepted and escrow is opening.'}
    if state == 'countered':
        return {'level': 'amber', 'headline': 'Status · Offer Countered',
                'message': 'A counter is out and we are awaiting the buyer response.'}
    if state == 'received':
        return {'level': 'amber', 'headline': 'Status · Offer Received',
                'message': 'An offer is in hand and awaiting your decision: accept, counter, or decline.'}
    this_week = weekly['this_week']
    tw_showings = this_week['showings'] if this_week else 0
    warm = len(detect_warm_leads(data))
    if weekly['momentum'] == 'accelerating' and tw_showings >= 1:
        return {'level': 'green', 'headline': 'Status · Active',
                'message': 'Activity is picking up week over week, warm leads in motion.'}
    if tw_showings == 0 and warm < 2:
        return {'level': 'amber', 'headline': 'Status · Attention Needed',
                'message': 'Activity has cooled this week. A decision on next phase is due.'}
    return {'level': 'amber', 'headline': 'Status · Watch',
            'message': 'Some activity, no offers yet.'}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--showing-file', required=True, type=Path)
    ap.add_argument('--address', required=True)
    ap.add_argument('--city', required=True)
    ap.add_argument('--state', default='CA')
    ap.add_argument('--beds', type=int, required=True)
    ap.add_argument('--baths', type=float, required=True)
    ap.add_argument('--sqft', type=int, required=True)
    ap.add_argument('--list-price', type=int, required=True)
    ap.add_argument('--mls', required=True)
    ap.add_argument('--seller-name', required=True)
    ap.add_argument('--listed-date', required=True)
    ap.add_argument('--report-date', default=None)
    ap.add_argument('--recent-weeks', type=int, default=6,
                    help='How many most-recent weeks to show individually before rolling older ones into "Earlier".')
    ap.add_argument('--offers', default=None,
                    help='Optional JSON array of confirmed offers, e.g. '
                         '[{"agent":"Sheila S.","price":1050000,"status":"accepted","terms":"...","date":"2026-04-15"}]')
    ap.add_argument('--output', required=True, type=Path)
    args = ap.parse_args()

    report_date = datetime.now() if not args.report_date else datetime.strptime(args.report_date, '%Y-%m-%d')
    listed_date = datetime.strptime(args.listed_date, '%Y-%m-%d')
    dom = (report_date - listed_date).days

    data = parse_showing_file(args.showing_file, default_year=report_date.year)
    weekly = build_weekly(data['all_records'], listed_date, report_date, recent_weeks=args.recent_weeks)
    chart_svg = build_chart_svg(weekly['weeks'])

    # offers: explicit --offers wins; otherwise fall back to candidate signals
    offer_signals = detect_offer_signals(data)
    offers = []
    if args.offers:
        try:
            offers = json.loads(args.offers)
        except json.JSONDecodeError:
            print("WARN: --offers was not valid JSON; using detected signals instead.", file=sys.stderr)
    offers_count = len(offers)
    offer_state = top_offer_state(offers)

    warm_leads = detect_warm_leads(data)
    status = determine_status(data, weekly, offers)
    ppsf = round(args.list_price / args.sqft)

    # dominant themes overall, sorted
    sorted_themes = sorted(
        [(THEME_LABELS.get(k, k), v) for k, v in data['themes'].items() if v > 0],
        key=lambda kv: kv[1], reverse=True
    )

    out = {
        'address': args.address,
        'report_date': report_date.strftime('%Y-%m-%d'),
        'dom': dom,
        'ppsf': ppsf,
        'status': status,
        'momentum': weekly['momentum'],
        'cumulative': weekly['cumulative'],
        'this_week': weekly['this_week'],
        'this_week_deltas': weekly['this_week_deltas'],
        'weeks': weekly['weeks'],
        'detailed_weeks': weekly['detailed_weeks'],
        'earlier_rollup': weekly['earlier_rollup'],
        'chart_svg': chart_svg,
        'undated': weekly['undated'],
        'dated': weekly['dated'],
        'offers': offers,
        'offer_signals': offer_signals,
        'offers_count': offers_count,
        'offer_state': offer_state,
        'warm_leads': [l['name'] for l in warm_leads],
        'themes_overall': sorted_themes,
        'counts': data['counts'],
        'records': {
            'showings': data['supra_agents'],
            'disclosures': data['glide_agents'],
            'open_house': data['openhouse_agents'],
        },
    }
    print(json.dumps(out, indent=2, default=str))

    if not weekly['dated']:
        print("\nNOTE: No parseable dates found in the export, so the week-over-week "
              "curve could not be built. The report will fall back to a single 'this period' "
              "view. Check that the export includes Date/Time columns.", file=sys.stderr)


if __name__ == '__main__':
    main()
